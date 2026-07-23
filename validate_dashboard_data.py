from __future__ import annotations

import json
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DATA_PATH = ROOT / "audit-data.js"
DATA_PREFIX = "window.auditDashboardData="

RECONCILED_STATUSES = {
    "Automatically Reconciled",
    "Accepted with user overrides",
    "Accepted with no user overrides",
}
EXCEPTION_STATUSES = {
    "Maintained in Exceptions with user overrides",
    "Maintained in Exceptions with no user overrides",
}
SOURCE_COLORS = {
    "Volume-adjusted / scaled blend": "#0b7f82",
    "Mixed Forecast Source": "#2f6fed",
    "100% Hybrid Forecast": "#e49a3a",
    "100% Final Forecast": "#99a9b8",
    "No active blend weight": "#b55f62",
}
SOURCE_METRICS = {
    "100% Hybrid Forecast": (
        "Exceptions Items Driven 100% by Hybrid Forecast",
        "Reconciled Items Driven 100% by Hybrid Forecast",
    ),
    "100% Final Forecast": (
        "Exceptions Items Driven 100% by Final Forecast",
        "Reconciled Items With 100% Final Forecast",
    ),
    "Mixed Forecast Source": (
        "Exceptions Items With Mixed Forecast Source",
        "Reconciled Items With Mixed Forecast Source",
    ),
    "Volume-adjusted / scaled blend": (
        "Exceptions Items With Volume-Adjusted / Scaled Blend",
        "Reconciled Items With Volume-Adjusted / Scaled Blend",
    ),
    "No active blend weight": (
        "Exceptions Items With No Active Blend Weight",
        "Reconciled Items With No Active Blend Weight",
    ),
}
RECONCILIATION_OVERRIDE_METRICS = {
    "automatic": (
        "Automatically Reconciled Items With User Overrides",
        "Automatically Reconciled Items Without User Overrides",
    ),
    "manual": (
        "Manually Reconciled Items With User Overrides",
        "Manually Reconciled Items Without User Overrides",
    ),
}


def canonical_source(value: object) -> str:
    label = str(value)
    if label.lower() == "volume-adjusted / scaled blend":
        return "Volume-adjusted / scaled blend"
    return label


def parse_target_month(value: object) -> datetime:
    if isinstance(value, (date, datetime)):
        return datetime(value.year, value.month, 1)
    return datetime.strptime(str(value).strip(), "%b-%y")


def find_row(rows: list[tuple], label: str) -> int:
    for index, row in enumerate(rows):
        if row and row[0] == label:
            return index
    raise AssertionError(f"Missing KPI row: {label}")


def section(rows: list[tuple], start: str, end: str | None) -> list[tuple]:
    start_index = find_row(rows, start) + 1
    end_index = find_row(rows, end) if end else len(rows)
    return rows[start_index:end_index]


def keyed_rows(rows: list[tuple]) -> dict[str, tuple]:
    return {
        str(row[0]): row
        for row in rows
        if row and row[0] is not None and row[1] is not None
    }


def as_int(value: object) -> int:
    return int(value or 0)


def assert_close(actual: float, expected: float, label: str) -> None:
    if abs(actual - expected) > 0.00011:
        raise AssertionError(f"{label}: actual={actual}, expected={expected}")


def load_payload() -> dict:
    text = DATA_PATH.read_text(encoding="utf-8").strip()
    payload_start = text.find(DATA_PREFIX)
    if payload_start < 0 or not text.endswith(";"):
        raise AssertionError("audit-data.js has an unexpected wrapper")
    return json.loads(text[payload_start + len(DATA_PREFIX) : -1])


def build_scope_kpi(metrics: dict[str, object]) -> dict:
    original_exceptions = metrics.get("Original Exception Items Count")
    reconciliation_overrides = None
    required_override_metrics = {
        metric
        for metric_pair in RECONCILIATION_OVERRIDE_METRICS.values()
        for metric in metric_pair
    }
    if required_override_metrics.issubset(metrics):
        reconciliation_overrides = {
            method: {
                "withOverrides": as_int(metrics[with_metric]),
                "withoutOverrides": as_int(metrics[without_metric]),
            }
            for method, (with_metric, without_metric) in (
                RECONCILIATION_OVERRIDE_METRICS.items()
            )
        }
    planner = {
        "reconciled": {
            "withOverrides": as_int(
                metrics["Reconciled Items With Planner Adjusted Forecast"]
            ),
            "withoutOverrides": as_int(
                metrics["Reconciled Items Without Planner Adjusted Forecast"]
            ),
        },
        "exceptions": {
            "withOverrides": as_int(
                metrics["Exception Items With Planner Adjusted Forecast"]
            ),
            "withoutOverrides": as_int(
                metrics["Exception Items Without Planner Adjusted Forecast"]
            ),
        },
    }
    sources = []
    for label, (exception_metric, reconciled_metric) in SOURCE_METRICS.items():
        exceptions = as_int(metrics[exception_metric])
        reconciled = as_int(metrics[reconciled_metric])
        sources.append(
            {
                "label": label,
                "exceptions": exceptions,
                "reconciled": reconciled,
                "count": exceptions + reconciled,
                "color": SOURCE_COLORS[label],
            }
        )

    return {
        "total": as_int(metrics["Total Items Count"]),
        "reconciled": as_int(metrics["Total Combinations in Reconciled"]),
        "exceptions": as_int(metrics["Total Items in Exceptions"]),
        "automatic": as_int(metrics["Automatically Reconciled Items Count"]),
        "manual": as_int(metrics["Manually Reconciled Items Count"]),
        "originalExceptions": (
            as_int(original_exceptions) if original_exceptions is not None else None
        ),
        "reconciliationOverrides": reconciliation_overrides,
        "planner": planner,
        "overrides": (
            planner["reconciled"]["withOverrides"]
            + planner["exceptions"]["withOverrides"]
        ),
        "sources": sources,
    }


def load_class_kpis(workbook, target_month: datetime, filename: str) -> tuple[dict, dict]:
    if "KPI by ABC Class" not in workbook.sheetnames:
        return {}, {}

    rows = list(workbook["KPI by ABC Class"].iter_rows(values_only=True))
    expected_header = (
        "target_month",
        "emr_abc_class",
        "number",
        "group",
        "count",
    )
    if not rows or tuple(rows[0][:5]) != expected_header:
        raise AssertionError(f"{filename}: invalid KPI by ABC Class header")

    scope_rows: dict[str, list[tuple]] = defaultdict(list)
    for row in rows[1:]:
        if row[0] is None:
            continue
        if parse_target_month(row[0]) != target_month:
            raise AssertionError(
                f"{filename}: ABC KPI target month {row[0]} does not match "
                f"{target_month:%b-%y}"
            )
        scope_rows[str(row[1])].append(tuple(row[:5]))

    if "All ABC Classes" not in scope_rows:
        raise AssertionError(f"{filename}: missing All ABC Classes KPI scope")

    expected_metrics = {str(row[3]) for row in scope_rows["All ABC Classes"]}
    metric_count = max(as_int(row[2]) for row in scope_rows["All ABC Classes"])
    expected_numbers = list(range(1, metric_count + 1))
    for scope, rows_for_scope in scope_rows.items():
        numbers = [as_int(row[2]) for row in rows_for_scope]
        metrics = [str(row[3]) for row in rows_for_scope]
        if numbers != expected_numbers:
            raise AssertionError(
                f"{filename}: {scope} metric numbers are not exactly 1 through {metric_count}"
            )
        if len(metrics) != len(set(metrics)) or set(metrics) != expected_metrics:
            raise AssertionError(f"{filename}: {scope} has an incomplete KPI metric set")

    parsed = {
        scope: build_scope_kpi({str(row[3]): row[4] for row in rows_for_scope})
        for scope, rows_for_scope in scope_rows.items()
    }
    all_scope = parsed.pop("All ABC Classes")
    return all_scope, parsed


def validate_workbook(
    path: Path,
    payload_months: dict[str, dict],
    payload_parts_by_month: dict[str, list[list]],
) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        expected_sheets = ["KPIs", "Part Level Breakdown"]
        if "KPI by ABC Class" in workbook.sheetnames:
            expected_sheets = ["KPIs", "KPI by ABC Class", "Part Level Breakdown"]
        if workbook.sheetnames != expected_sheets:
            raise AssertionError(f"{path.name}: unexpected sheets {workbook.sheetnames}")

        kpi_rows = list(workbook["KPIs"].iter_rows(values_only=True))
        part_sheet = workbook["Part Level Breakdown"]
        header = tuple(next(part_sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        expected_header = (
            "item",
            "item_description",
            "emr_class",
            "target_month",
            "month_status",
            "forecast_source_classification",
        )
        if header[:6] != expected_header:
            raise AssertionError(f"{path.name}: unexpected Part Level header {header[:6]}")

        raw_parts = [
            tuple(row[:6])
            for row in part_sheet.iter_rows(min_row=2, values_only=True)
            if row[0] is not None
        ]
        if any(value is None for row in raw_parts for value in row):
            raise AssertionError(f"{path.name}: blank Part Level value found")

        item_counts = Counter(str(row[0]) for row in raw_parts)
        duplicate_items = [item for item, count in item_counts.items() if count > 1]
        if duplicate_items:
            raise AssertionError(f"{path.name}: duplicate items {duplicate_items[:10]}")

        target_months = {parse_target_month(row[3]) for row in raw_parts}
        if len(target_months) != 1:
            raise AssertionError(f"{path.name}: target months {sorted(target_months)}")
        target_month = target_months.pop()
        month_key = target_month.strftime("%Y-%m")
        if month_key not in payload_months or month_key not in payload_parts_by_month:
            raise AssertionError(f"{path.name}: month {month_key} missing from audit-data.js")
        payload_month = payload_months[month_key]
        payload_parts = payload_parts_by_month[month_key]

        expected_parts = [
            [
                str(row[0]),
                str(row[1]),
                str(row[2]),
                str(row[4]),
                canonical_source(row[5]),
                "with user overrides" in str(row[4]).lower(),
            ]
            for row in raw_parts
        ]
        if payload_parts != expected_parts:
            mismatch = next(
                (
                    index
                    for index, (actual, expected) in enumerate(
                        zip(payload_parts, expected_parts, strict=False)
                    )
                    if actual != expected
                ),
                min(len(payload_parts), len(expected_parts)),
            )
            raise AssertionError(f"{path.name}: serialized Part Level mismatch at row {mismatch + 2}")

        overall = keyed_rows(
            section(kpi_rows, "1. Overall Status Summary", "2. Reconciliation Type")
        )
        reconciliation = keyed_rows(
            section(kpi_rows, "2. Reconciliation Type", "3. Planner Adjusted Forecast")
        )
        planner = keyed_rows(
            section(kpi_rows, "3. Planner Adjusted Forecast", "4. Forecast Source Distribution")
        )
        source_rows = keyed_rows(
            section(kpi_rows, "4. Forecast Source Distribution", "5. ABC Class Breakdown")
        )
        abc_rows = keyed_rows(section(kpi_rows, "5. ABC Class Breakdown", None))

        total = as_int(overall["Total Items"][1])
        reconciled = as_int(overall["Reconciled"][1])
        exceptions = as_int(overall["Exceptions"][1])
        automatic = as_int(reconciliation["Automatically Reconciled"][1])
        manual = as_int(reconciliation["Manually Reconciled"][1])
        planner_data = {
            "reconciled": {
                "withOverrides": as_int(planner["Reconciled"][1]),
                "withoutOverrides": as_int(planner["Reconciled"][2]),
            },
            "exceptions": {
                "withOverrides": as_int(planner["Exceptions"][1]),
                "withoutOverrides": as_int(planner["Exceptions"][2]),
            },
        }
        sources = [
            {
                "label": canonical_source(label),
                "exceptions": as_int(row[1]),
                "reconciled": as_int(row[2]),
                "count": as_int(row[3]),
                "color": SOURCE_COLORS[canonical_source(label)],
            }
            for label, row in source_rows.items()
            if label not in {"Forecast Source", "Total"}
        ]
        abc = [
            {
                "label": label,
                "total": as_int(row[1]),
                "reconciled": as_int(row[2]),
                "exceptions": as_int(row[3]),
            }
            for label, row in abc_rows.items()
            if label != "emr_abc_class"
        ]
        all_class_scope, class_kpis = load_class_kpis(
            workbook,
            target_month,
            path.name,
        )
        expected_month = {
            "key": month_key,
            "label": target_month.strftime("%b"),
            "year": target_month.strftime("%Y"),
            "total": total,
            "reconciled": reconciled,
            "exceptions": exceptions,
            "automatic": automatic,
            "manual": manual,
            "originalExceptions": all_class_scope.get("originalExceptions"),
            "reconciliationOverrides": all_class_scope.get(
                "reconciliationOverrides"
            ),
            "planner": planner_data,
            "overrides": (
                planner_data["reconciled"]["withOverrides"]
                + planner_data["exceptions"]["withOverrides"]
            ),
            "sources": sources,
            "abc": abc,
            "kpiByAbc": class_kpis,
        }
        if payload_month != expected_month:
            raise AssertionError(f"{path.name}: serialized KPI payload differs from workbook")

        if class_kpis:
            scopes = {"All ABC Classes": all_class_scope, **class_kpis}
            for label, scope in scopes.items():
                if scope["originalExceptions"] is not None:
                    expected_original = scope["exceptions"] + scope["manual"]
                    if scope["originalExceptions"] != expected_original:
                        raise AssertionError(
                            f"{path.name}: {label} original exceptions "
                            f"{scope['originalExceptions']} != {expected_original}"
                        )
                    if scope["originalExceptions"] != scope["total"] - scope["automatic"]:
                        raise AssertionError(
                            f"{path.name}: {label} original exceptions do not "
                            "exclude only automatically reconciled items"
                        )

                reconciliation_overrides = scope["reconciliationOverrides"]
                if reconciliation_overrides:
                    automatic_split = reconciliation_overrides["automatic"]
                    manual_split = reconciliation_overrides["manual"]
                    split_checks = {
                        "automatic": (
                            automatic_split["withOverrides"]
                            + automatic_split["withoutOverrides"],
                            scope["automatic"],
                        ),
                        "manual": (
                            manual_split["withOverrides"]
                            + manual_split["withoutOverrides"],
                            scope["manual"],
                        ),
                        "with override": (
                            automatic_split["withOverrides"]
                            + manual_split["withOverrides"],
                            scope["planner"]["reconciled"]["withOverrides"],
                        ),
                        "without override": (
                            automatic_split["withoutOverrides"]
                            + manual_split["withoutOverrides"],
                            scope["planner"]["reconciled"]["withoutOverrides"],
                        ),
                    }
                    for split_label, (actual, expected) in split_checks.items():
                        if actual != expected:
                            raise AssertionError(
                                f"{path.name}: {label} reconciliation "
                                f"{split_label} {actual} != {expected}"
                            )

            reconciliation_overrides = all_class_scope["reconciliationOverrides"]
            if reconciliation_overrides:
                status_counts = Counter(part[3] for part in payload_parts)
                manual_split = reconciliation_overrides["manual"]
                manual_status_checks = {
                    "manual with override": (
                        manual_split["withOverrides"],
                        status_counts["Accepted with user overrides"],
                    ),
                    "manual without override": (
                        manual_split["withoutOverrides"],
                        status_counts["Accepted with no user overrides"],
                    ),
                }
                for label, (actual, expected) in manual_status_checks.items():
                    if actual != expected:
                        raise AssertionError(
                            f"{path.name}: {label} {actual} != {expected}"
                        )

            class_checks = {
                "total": total,
                "reconciled": reconciled,
                "exceptions": exceptions,
                "automatic": automatic,
                "manual": manual,
                "overrides": expected_month["overrides"],
            }
            if expected_month["originalExceptions"] is not None:
                class_checks["originalExceptions"] = expected_month[
                    "originalExceptions"
                ]
            for field, expected in class_checks.items():
                if all_class_scope[field] != expected:
                    raise AssertionError(
                        f"{path.name}: All ABC Classes {field} "
                        f"{all_class_scope[field]} != {expected}"
                    )
                class_sum = sum(scope[field] for scope in class_kpis.values())
                if class_sum != expected:
                    raise AssertionError(
                        f"{path.name}: ABC class {field} sum {class_sum} != {expected}"
                    )

            if all_class_scope["reconciliationOverrides"]:
                for method in ("automatic", "manual"):
                    for field in ("withOverrides", "withoutOverrides"):
                        expected = all_class_scope["reconciliationOverrides"][
                            method
                        ][field]
                        class_sum = sum(
                            scope["reconciliationOverrides"][method][field]
                            for scope in class_kpis.values()
                        )
                        if class_sum != expected:
                            raise AssertionError(
                                f"{path.name}: ABC class {method} {field} "
                                f"sum {class_sum} != {expected}"
                            )

            class_totals = {
                label: scope["total"] for label, scope in class_kpis.items()
            }
            workbook_abc_totals = {row["label"]: row["total"] for row in abc}
            if class_totals != workbook_abc_totals:
                raise AssertionError(
                    f"{path.name}: filtered KPI class totals differ from ABC breakdown"
                )

            all_sources = {
                source["label"]: source["count"]
                for source in all_class_scope["sources"]
            }
            for source in sources:
                if all_sources[source["label"]] != source["count"]:
                    raise AssertionError(
                        f"{path.name}: All ABC Classes source {source['label']} differs"
                    )
                class_source_sum = sum(
                    next(
                        row["count"]
                        for row in scope["sources"]
                        if row["label"] == source["label"]
                    )
                    for scope in class_kpis.values()
                )
                if class_source_sum != source["count"]:
                    raise AssertionError(
                        f"{path.name}: ABC class source {source['label']} sum differs"
                    )

        statuses = Counter(str(row[4]) for row in raw_parts)
        unknown_statuses = set(statuses) - RECONCILED_STATUSES - EXCEPTION_STATUSES
        if unknown_statuses:
            raise AssertionError(f"{path.name}: unknown statuses {sorted(unknown_statuses)}")

        source_splits = Counter()
        abc_splits = Counter()
        for row in raw_parts:
            cohort = "reconciled" if row[4] in RECONCILED_STATUSES else "exceptions"
            source_splits[(canonical_source(row[5]), cohort)] += 1
            abc_splits[(str(row[2]), cohort)] += 1

        expected_source_labels = {source["label"] for source in sources}
        actual_source_labels = {canonical_source(row[5]) for row in raw_parts}
        if not actual_source_labels.issubset(expected_source_labels):
            raise AssertionError(
                f"{path.name}: Part Level sources missing from KPI sheet "
                f"{sorted(actual_source_labels - expected_source_labels)}"
            )
        expected_abc_labels = {row["label"] for row in abc}
        actual_abc_labels = {str(row[2]) for row in raw_parts}
        if actual_abc_labels != expected_abc_labels:
            raise AssertionError(
                f"{path.name}: ABC labels differ: "
                f"Part Level={sorted(actual_abc_labels)}, KPI={sorted(expected_abc_labels)}"
            )

        checks = {
            "total": (len(raw_parts), total),
            "overall split": (reconciled + exceptions, total),
            "reconciled": (
                sum(statuses[status] for status in RECONCILED_STATUSES),
                reconciled,
            ),
            "exceptions": (
                sum(statuses[status] for status in EXCEPTION_STATUSES),
                exceptions,
            ),
            "automatic": (statuses["Automatically Reconciled"], automatic),
            "manual": (
                statuses["Accepted with user overrides"]
                + statuses["Accepted with no user overrides"],
                manual,
            ),
            "reconciliation total": (
                as_int(reconciliation["Total Reconciled"][1]),
                reconciled,
            ),
            "reconciled planner total": (
                planner_data["reconciled"]["withOverrides"]
                + planner_data["reconciled"]["withoutOverrides"],
                reconciled,
            ),
            "exception planner total": (
                planner_data["exceptions"]["withOverrides"]
                + planner_data["exceptions"]["withoutOverrides"],
                exceptions,
            ),
            "source KPI exceptions total": (
                as_int(source_rows["Total"][1]),
                exceptions,
            ),
            "source KPI reconciled total": (
                as_int(source_rows["Total"][2]),
                reconciled,
            ),
            "source KPI overall total": (
                as_int(source_rows["Total"][3]),
                total,
            ),
        }
        for source in sources:
            checks[f"{source['label']} reconciled"] = (
                source_splits[(source["label"], "reconciled")],
                source["reconciled"],
            )
            checks[f"{source['label']} exceptions"] = (
                source_splits[(source["label"], "exceptions")],
                source["exceptions"],
            )
            checks[f"{source['label']} total"] = (
                source_splits[(source["label"], "reconciled")]
                + source_splits[(source["label"], "exceptions")],
                source["count"],
            )
        for row in abc:
            checks[f"ABC {row['label']} reconciled"] = (
                abc_splits[(row["label"], "reconciled")],
                row["reconciled"],
            )
            checks[f"ABC {row['label']} exceptions"] = (
                abc_splits[(row["label"], "exceptions")],
                row["exceptions"],
            )
            checks[f"ABC {row['label']} total"] = (
                abc_splits[(row["label"], "reconciled")]
                + abc_splits[(row["label"], "exceptions")],
                row["total"],
            )
        failures = [
            f"{label}: actual={actual}, expected={expected}"
            for label, (actual, expected) in checks.items()
            if actual != expected
        ]
        if failures:
            raise AssertionError(f"{path.name}: " + "; ".join(failures))

        assert_close(float(overall["Reconciled"][2]), reconciled / total, f"{path.name} reconciled share")
        assert_close(float(overall["Exceptions"][2]), exceptions / total, f"{path.name} exception share")
        assert_close(
            float(reconciliation["Automatically Reconciled"][2]),
            automatic / reconciled,
            f"{path.name} automatic share",
        )
        assert_close(
            float(reconciliation["Manually Reconciled"][2]),
            manual / reconciled,
            f"{path.name} manual share",
        )
        assert_close(
            float(planner["Reconciled"][4]),
            planner_data["reconciled"]["withOverrides"] / reconciled,
            f"{path.name} reconciled override share",
        )
        assert_close(
            float(planner["Exceptions"][4]),
            planner_data["exceptions"]["withOverrides"] / exceptions,
            f"{path.name} exception override share",
        )
        for row in abc_rows.values():
            if row[0] == "emr_abc_class":
                continue
            expected_rate = as_int(row[2]) / as_int(row[1]) if as_int(row[1]) else 0
            assert_close(float(row[4]), expected_rate, f"{path.name} ABC {row[0]} rate")

        part_override_count = (
            statuses["Accepted with user overrides"]
            + statuses["Maintained in Exceptions with user overrides"]
        )
        return {
            "key": month_key,
            "rows": len(raw_parts),
            "items": set(item_counts),
            "parts": expected_parts,
            "partOverrideCount": part_override_count,
            "kpiOverrideCount": expected_month["overrides"],
        }
    finally:
        workbook.close()


def activity_metrics(parts_by_month: dict[str, list[list]], month_keys: list[str]) -> dict:
    metrics: dict[str, dict[str, int]] = defaultdict(
        lambda: {"automatic": 0, "adjusted": 0, "manuallyAccepted": 0}
    )
    for key in month_keys[-6:]:
        for part in parts_by_month[key]:
            item_metrics = metrics[part[0]]
            item_metrics["automatic"] += part[3] == "Automatically Reconciled"
            item_metrics["adjusted"] += bool(part[5])
            item_metrics["manuallyAccepted"] += part[3].startswith("Accepted with")
    return dict(metrics)


def main() -> None:
    payload = load_payload()
    workbook_paths = sorted(ROOT.glob("EMR - Audit Reports - *.xlsx"))
    if not workbook_paths:
        raise SystemExit("No audit workbooks found")

    payload_months = {month["key"]: month for month in payload["months"]}
    workbook_results = []
    expected_keys = []
    for path in workbook_paths:
        result = validate_workbook(
            path,
            payload_months,
            payload["partsByMonth"],
        )
        expected_keys.append(result["key"])
        workbook_results.append(result)

    expected_keys.sort()
    payload_keys = [month["key"] for month in payload["months"]]
    if payload_keys != expected_keys:
        raise AssertionError(f"Month order mismatch: {payload_keys} != {expected_keys}")
    if set(payload["partsByMonth"]) != set(expected_keys):
        raise AssertionError("audit-data.js has missing or extra part-level months")

    source_metrics = activity_metrics(payload["partsByMonth"], expected_keys)
    independent_parts = {
        result["key"]: result["parts"]
        for result in workbook_results
    }
    independent_metrics = activity_metrics(independent_parts, expected_keys)
    if source_metrics != independent_metrics:
        raise AssertionError("Six-month part activity metrics do not match source rows")

    all_items = set().union(*(result["items"] for result in workbook_results))
    if set(source_metrics) != all_items:
        raise AssertionError("Six-month metrics do not cover every workbook item")

    print("Dashboard data validation: PASS")
    print(
        f"Validated {len(workbook_results)} workbooks, "
        f"{sum(result['rows'] for result in workbook_results):,} part rows, "
        f"and {len(all_items):,} unique items."
    )
    for result in sorted(workbook_results, key=lambda value: value["key"]):
        gap = result["kpiOverrideCount"] - result["partOverrideCount"]
        print(
            f"{result['key']}: {result['rows']:,} rows; "
            f"KPI overrides={result['kpiOverrideCount']:,}; "
            f"override-tagged part rows={result['partOverrideCount']:,}; gap={gap:,}"
        )
    print(
        "Note: the Part Level sheet does not identify overrides on rows classified "
        "as Automatically Reconciled. The per-item Adjusted metric therefore counts "
        "only rows whose part-level status explicitly says 'with user overrides'."
    )


if __name__ == "__main__":
    main()
