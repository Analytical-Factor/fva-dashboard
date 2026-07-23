from __future__ import annotations

import json
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
OUTPUT_PATH = ROOT / "audit-data.js"
SOURCE_COLORS = {
    "Volume-adjusted / scaled blend": "#0b7f82",
    "Mixed Forecast Source": "#2f6fed",
    "100% Hybrid Forecast": "#e49a3a",
    "100% Final Forecast": "#99a9b8",
    "No active blend weight": "#b55f62",
}
SOURCE_METRICS = {
    "100% Hybrid Forecast": (
        "Exceptions Items Driven 100% by Hybrid Forecast",
        "Reconciled Items Driven 100% by Hybrid Forecast",
    ),
    "100% Final Forecast": (
        "Exceptions Items Driven 100% by Final Forecast",
        "Reconciled Items With 100% Final Forecast",
    ),
    "Mixed Forecast Source": (
        "Exceptions Items With Mixed Forecast Source",
        "Reconciled Items With Mixed Forecast Source",
    ),
    "Volume-adjusted / scaled blend": (
        "Exceptions Items With Volume-Adjusted / Scaled Blend",
        "Reconciled Items With Volume-Adjusted / Scaled Blend",
    ),
    "No active blend weight": (
        "Exceptions Items With No Active Blend Weight",
        "Reconciled Items With No Active Blend Weight",
    ),
}

RECONCILED_STATUSES = {
    "Automatically Reconciled",
    "Accepted with user overrides",
    "Accepted with no user overrides",
}
EXCEPTION_STATUSES = {
    "Maintained in Exceptions with user overrides",
    "Maintained in Exceptions with no user overrides",
}
RECONCILIATION_OVERRIDE_METRICS = {
    "automatic": (
        "Automatically Reconciled Items With User Overrides",
        "Automatically Reconciled Items Without User Overrides",
    ),
    "manual": (
        "Manually Reconciled Items With User Overrides",
        "Manually Reconciled Items Without User Overrides",
    ),
}


def find_row(rows: list[tuple], label: str) -> int:
    for index, row in enumerate(rows):
        if row and row[0] == label:
            return index
    raise ValueError(f"Missing KPI section or row: {label}")


def section(rows: list[tuple], start_label: str, end_label: str | None) -> list[tuple]:
    start = find_row(rows, start_label) + 1
    end = find_row(rows, end_label) if end_label else len(rows)
    return rows[start:end]


def keyed_rows(rows: list[tuple]) -> dict[str, tuple]:
    return {
        str(row[0]): row
        for row in rows
        if row and row[0] is not None and row[1] is not None
    }


def as_int(value: object) -> int:
    return int(value or 0)


def canonical_source(value: object) -> str:
    label = str(value)
    if label.lower() == "volume-adjusted / scaled blend":
        return "Volume-adjusted / scaled blend"
    return label


def parse_target_month(value: object) -> datetime:
    if isinstance(value, (date, datetime)):
        return datetime(value.year, value.month, 1)
    return datetime.strptime(str(value).strip(), "%b-%y")


def build_scope_kpi(metrics: dict[str, object]) -> dict:
    original_exceptions = metrics.get("Original Exception Items Count")
    reconciliation_overrides = None
    required_override_metrics = {
        metric
        for metric_pair in RECONCILIATION_OVERRIDE_METRICS.values()
        for metric in metric_pair
    }
    if required_override_metrics.issubset(metrics):
        reconciliation_overrides = {
            method: {
                "withOverrides": as_int(metrics[with_metric]),
                "withoutOverrides": as_int(metrics[without_metric]),
            }
            for method, (with_metric, without_metric) in (
                RECONCILIATION_OVERRIDE_METRICS.items()
            )
        }
    planner = {
        "reconciled": {
            "withOverrides": as_int(
                metrics["Reconciled Items With Planner Adjusted Forecast"]
            ),
            "withoutOverrides": as_int(
                metrics["Reconciled Items Without Planner Adjusted Forecast"]
            ),
        },
        "exceptions": {
            "withOverrides": as_int(
                metrics["Exception Items With Planner Adjusted Forecast"]
            ),
            "withoutOverrides": as_int(
                metrics["Exception Items Without Planner Adjusted Forecast"]
            ),
        },
    }
    sources = []
    for label, (exception_metric, reconciled_metric) in SOURCE_METRICS.items():
        exceptions = as_int(metrics[exception_metric])
        reconciled = as_int(metrics[reconciled_metric])
        sources.append(
            {
                "label": label,
                "exceptions": exceptions,
                "reconciled": reconciled,
                "count": exceptions + reconciled,
                "color": SOURCE_COLORS[label],
            }
        )

    return {
        "total": as_int(metrics["Total Items Count"]),
        "reconciled": as_int(metrics["Total Combinations in Reconciled"]),
        "exceptions": as_int(metrics["Total Items in Exceptions"]),
        "automatic": as_int(metrics["Automatically Reconciled Items Count"]),
        "manual": as_int(metrics["Manually Reconciled Items Count"]),
        "originalExceptions": (
            as_int(original_exceptions) if original_exceptions is not None else None
        ),
        "reconciliationOverrides": reconciliation_overrides,
        "planner": planner,
        "overrides": (
            planner["reconciled"]["withOverrides"]
            + planner["exceptions"]["withOverrides"]
        ),
        "sources": sources,
    }


def load_class_kpis(workbook, target_month: datetime, filename: str) -> tuple[dict, dict]:
    if "KPI by ABC Class" not in workbook.sheetnames:
        return {}, {}

    sheet = workbook["KPI by ABC Class"]
    rows = list(sheet.iter_rows(values_only=True))
    expected_header = (
        "target_month",
        "emr_abc_class",
        "number",
        "group",
        "count",
    )
    if not rows or tuple(rows[0][:5]) != expected_header:
        raise ValueError(f"{filename}: invalid KPI by ABC Class header")

    scope_metrics: dict[str, dict[str, object]] = {}
    sheet_months = set()
    for row in rows[1:]:
        if row[0] is None:
            continue
        sheet_months.add(parse_target_month(row[0]))
        scope = str(row[1])
        scope_metrics.setdefault(scope, {})[str(row[3])] = row[4]

    if sheet_months != {target_month}:
        raise ValueError(
            f"{filename}: KPI by ABC Class target month {sheet_months} "
            f"does not match {target_month:%b-%y}"
        )
    if "All ABC Classes" not in scope_metrics:
        raise ValueError(f"{filename}: missing All ABC Classes KPI scope")

    all_scope = build_scope_kpi(scope_metrics["All ABC Classes"])
    class_kpis = {
        scope: build_scope_kpi(metrics)
        for scope, metrics in scope_metrics.items()
        if scope != "All ABC Classes"
    }
    return all_scope, class_kpis


def load_workbook_data(path: Path) -> tuple[dict, list[list]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        kpi_rows = list(workbook["KPIs"].iter_rows(values_only=True))
        part_sheet = workbook["Part Level Breakdown"]

        overall = keyed_rows(
            section(kpi_rows, "1. Overall Status Summary", "2. Reconciliation Type")
        )
        reconciliation = keyed_rows(
            section(kpi_rows, "2. Reconciliation Type", "3. Planner Adjusted Forecast")
        )
        planner = keyed_rows(
            section(kpi_rows, "3. Planner Adjusted Forecast", "4. Forecast Source Distribution")
        )
        source_rows = keyed_rows(
            section(kpi_rows, "4. Forecast Source Distribution", "5. ABC Class Breakdown")
        )
        abc_rows = keyed_rows(section(kpi_rows, "5. ABC Class Breakdown", None))

        parts: list[list] = []
        target_months: set[datetime] = set()
        for row in part_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            item, description, abc_class, target_month, status, source = row[:6]
            target_months.add(parse_target_month(target_month))
            parts.append(
                [
                    str(item),
                    str(description or ""),
                    str(abc_class or "Unclassified"),
                    str(status),
                    canonical_source(source),
                    "with user overrides" in str(status).lower(),
                ]
            )

        if len(target_months) != 1:
            raise ValueError(f"{path.name}: expected one target month, found {target_months}")

        target_month = target_months.pop()
        all_class_scope, class_kpis = load_class_kpis(
            workbook,
            target_month,
            path.name,
        )
        total = as_int(overall["Total Items"][1])
        reconciled = as_int(overall["Reconciled"][1])
        exceptions = as_int(overall["Exceptions"][1])
        automatic = as_int(reconciliation["Automatically Reconciled"][1])
        manual = as_int(reconciliation["Manually Reconciled"][1])

        planner_data = {
            "reconciled": {
                "withOverrides": as_int(planner["Reconciled"][1]),
                "withoutOverrides": as_int(planner["Reconciled"][2]),
            },
            "exceptions": {
                "withOverrides": as_int(planner["Exceptions"][1]),
                "withoutOverrides": as_int(planner["Exceptions"][2]),
            },
        }

        sources = []
        for label, row in source_rows.items():
            if label in {"Forecast Source", "Total"}:
                continue
            label = canonical_source(label)
            sources.append(
                {
                    "label": label,
                    "exceptions": as_int(row[1]),
                    "reconciled": as_int(row[2]),
                    "count": as_int(row[3]),
                    "color": SOURCE_COLORS.get(label, "#647985"),
                }
            )

        abc = []
        for label, row in abc_rows.items():
            if label == "emr_abc_class":
                continue
            abc.append(
                {
                    "label": label,
                    "total": as_int(row[1]),
                    "reconciled": as_int(row[2]),
                    "exceptions": as_int(row[3]),
                }
            )

        month = {
            "key": target_month.strftime("%Y-%m"),
            "label": target_month.strftime("%b"),
            "year": target_month.strftime("%Y"),
            "total": total,
            "reconciled": reconciled,
            "exceptions": exceptions,
            "automatic": automatic,
            "manual": manual,
            "originalExceptions": all_class_scope.get("originalExceptions"),
            "reconciliationOverrides": all_class_scope.get(
                "reconciliationOverrides"
            ),
            "planner": planner_data,
            "overrides": (
                planner_data["reconciled"]["withOverrides"]
                + planner_data["exceptions"]["withOverrides"]
            ),
            "sources": sources,
            "abc": abc,
            "kpiByAbc": class_kpis,
        }

        validate_month(path.name, month, parts)
        validate_class_kpis(path.name, month, all_class_scope)
        return month, parts
    finally:
        workbook.close()


def validate_month(filename: str, month: dict, parts: list[list]) -> None:
    status_counts = Counter(part[3] for part in parts)
    unknown_statuses = set(status_counts) - RECONCILED_STATUSES - EXCEPTION_STATUSES
    if unknown_statuses:
        raise ValueError(f"{filename}: unknown statuses: {sorted(unknown_statuses)}")

    reconciled_count = sum(status_counts[status] for status in RECONCILED_STATUSES)
    exception_count = sum(status_counts[status] for status in EXCEPTION_STATUSES)
    automatic_count = status_counts["Automatically Reconciled"]
    manual_count = reconciled_count - automatic_count
    source_counts = Counter(part[4] for part in parts)
    abc_total_counts = Counter(part[2] for part in parts)
    abc_reconciled_counts = Counter(
        part[2] for part in parts if part[3] in RECONCILED_STATUSES
    )

    checks = {
        "part rows": (len(parts), month["total"]),
        "reconciled rows": (reconciled_count, month["reconciled"]),
        "exception rows": (exception_count, month["exceptions"]),
        "automatic rows": (automatic_count, month["automatic"]),
        "manual rows": (manual_count, month["manual"]),
        "reconciled planner total": (
            month["planner"]["reconciled"]["withOverrides"]
            + month["planner"]["reconciled"]["withoutOverrides"],
            month["reconciled"],
        ),
        "exception planner total": (
            month["planner"]["exceptions"]["withOverrides"]
            + month["planner"]["exceptions"]["withoutOverrides"],
            month["exceptions"],
        ),
    }
    reconciliation_overrides = month["reconciliationOverrides"]
    if reconciliation_overrides:
        automatic = reconciliation_overrides["automatic"]
        manual = reconciliation_overrides["manual"]
        checks.update(
            {
                "automatic override split": (
                    automatic["withOverrides"] + automatic["withoutOverrides"],
                    month["automatic"],
                ),
                "manual override split": (
                    manual["withOverrides"] + manual["withoutOverrides"],
                    month["manual"],
                ),
                "reconciled with-override split": (
                    automatic["withOverrides"] + manual["withOverrides"],
                    month["planner"]["reconciled"]["withOverrides"],
                ),
                "reconciled without-override split": (
                    automatic["withoutOverrides"] + manual["withoutOverrides"],
                    month["planner"]["reconciled"]["withoutOverrides"],
                ),
                "manual with-override status rows": (
                    manual["withOverrides"],
                    status_counts["Accepted with user overrides"],
                ),
                "manual without-override status rows": (
                    manual["withoutOverrides"],
                    status_counts["Accepted with no user overrides"],
                ),
            }
        )

    for source in month["sources"]:
        checks[f"source {source['label']}"] = (
            source_counts[source["label"]],
            source["count"],
        )

    for row in month["abc"]:
        checks[f"ABC {row['label']} total"] = (
            abc_total_counts[row["label"]],
            row["total"],
        )
        checks[f"ABC {row['label']} reconciled"] = (
            abc_reconciled_counts[row["label"]],
            row["reconciled"],
        )

    failures = [
        f"{name}: actual={actual}, KPI={expected}"
        for name, (actual, expected) in checks.items()
        if actual != expected
    ]
    if failures:
        raise ValueError(f"{filename} validation failed: " + "; ".join(failures))


def validate_class_kpis(filename: str, month: dict, all_scope: dict) -> None:
    if not month["kpiByAbc"]:
        return

    aggregate_checks = {
        "total": month["total"],
        "reconciled": month["reconciled"],
        "exceptions": month["exceptions"],
        "automatic": month["automatic"],
        "manual": month["manual"],
        "overrides": month["overrides"],
    }
    if month["originalExceptions"] is not None:
        aggregate_checks["originalExceptions"] = month["originalExceptions"]
    failures = []
    scopes = {"All ABC Classes": all_scope, **month["kpiByAbc"]}
    for label, scope in scopes.items():
        if scope["originalExceptions"] is not None:
            expected_original = scope["exceptions"] + scope["manual"]
            if scope["originalExceptions"] != expected_original:
                failures.append(
                    f"{label} original exceptions: "
                    f"{scope['originalExceptions']} != {expected_original}"
                )
            if scope["originalExceptions"] != scope["total"] - scope["automatic"]:
                failures.append(
                    f"{label} original exceptions do not exclude only automatic items"
                )

        reconciliation_overrides = scope["reconciliationOverrides"]
        if reconciliation_overrides:
            automatic = reconciliation_overrides["automatic"]
            manual = reconciliation_overrides["manual"]
            split_checks = {
                "automatic": (
                    automatic["withOverrides"] + automatic["withoutOverrides"],
                    scope["automatic"],
                ),
                "manual": (
                    manual["withOverrides"] + manual["withoutOverrides"],
                    scope["manual"],
                ),
                "with override": (
                    automatic["withOverrides"] + manual["withOverrides"],
                    scope["planner"]["reconciled"]["withOverrides"],
                ),
                "without override": (
                    automatic["withoutOverrides"] + manual["withoutOverrides"],
                    scope["planner"]["reconciled"]["withoutOverrides"],
                ),
            }
            for split_label, (actual, expected) in split_checks.items():
                if actual != expected:
                    failures.append(
                        f"{label} reconciliation {split_label}: "
                        f"{actual} != {expected}"
                    )

    for field, expected in aggregate_checks.items():
        if all_scope[field] != expected:
            failures.append(
                f"All ABC Classes {field}: query={all_scope[field]}, KPI={expected}"
            )
        class_sum = sum(scope[field] for scope in month["kpiByAbc"].values())
        if class_sum != expected:
            failures.append(
                f"ABC class sum {field}: query={class_sum}, KPI={expected}"
            )

    if all_scope["reconciliationOverrides"]:
        for method in ("automatic", "manual"):
            for field in ("withOverrides", "withoutOverrides"):
                expected = all_scope["reconciliationOverrides"][method][field]
                class_sum = sum(
                    scope["reconciliationOverrides"][method][field]
                    for scope in month["kpiByAbc"].values()
                )
                if class_sum != expected:
                    failures.append(
                        f"ABC class {method} {field} sum: "
                        f"{class_sum} != {expected}"
                    )

    all_sources = {source["label"]: source for source in all_scope["sources"]}
    workbook_sources = {source["label"]: source for source in month["sources"]}
    for label, source in all_sources.items():
        workbook_source = workbook_sources.get(label)
        expected = workbook_source["count"] if workbook_source else 0
        if source["count"] != expected:
            failures.append(
                f"All ABC Classes source {label}: query={source['count']}, KPI={expected}"
            )
        class_sum = sum(
            next(
                row["count"]
                for row in scope["sources"]
                if row["label"] == label
            )
            for scope in month["kpiByAbc"].values()
        )
        if class_sum != source["count"]:
            failures.append(
                f"ABC class source sum {label}: {class_sum} != {source['count']}"
            )

    abc_totals = {row["label"]: row["total"] for row in month["abc"]}
    class_totals = {
        label: scope["total"] for label, scope in month["kpiByAbc"].items()
    }
    if class_totals != abc_totals:
        failures.append(
            f"ABC totals differ: class KPI={class_totals}, KPI sheet={abc_totals}"
        )

    if failures:
        raise ValueError(
            f"{filename} KPI by ABC Class validation failed: " + "; ".join(failures)
        )


def main() -> None:
    workbook_paths = sorted(ROOT.glob("EMR - Audit Reports - *.xlsx"))
    if not workbook_paths:
        raise SystemExit("No audit report workbooks found.")

    months = []
    parts_by_month = {}
    for workbook_path in workbook_paths:
        month, parts = load_workbook_data(workbook_path)
        months.append(month)
        parts_by_month[month["key"]] = parts

    months.sort(key=lambda month: month["key"])
    data = {
        "months": months,
        "partsByMonth": {month["key"]: parts_by_month[month["key"]] for month in months},
    }

    payload = json.dumps(data, ensure_ascii=True, separators=(",", ":"))
    OUTPUT_PATH.write_text(
        "// Generated from the audit workbooks by build_audit_data.py.\n"
        f"window.auditDashboardData={payload};\n",
        encoding="utf-8",
    )

    print(f"Generated {OUTPUT_PATH.name} from {len(months)} workbooks.")
    for month in months:
        print(
            f"{month['label']} {month['year']}: "
            f"{month['total']:,} items, {month['reconciled']:,} reconciled, "
            f"{len(parts_by_month[month['key']]):,} part rows, "
            f"{len(month['kpiByAbc'])} filterable ABC classes"
        )


if __name__ == "__main__":
    main()
